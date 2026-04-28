import base64
import io
import json
import time

import boto3
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REQUEST_TIMEOUT = 25
DISCOVER_DELAY = 0.05

HUBSPOT_BASE = "https://api.hubapi.com"
HUBSPOT_STANDARD_OBJECTS = [
    "contacts", "companies", "deals", "tickets",
    "line_items", "products", "quotes", "calls",
    "emails", "meetings", "notes", "tasks", "communications",
]

CORS = {
    "Content-Type": "application/json",
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Headers": "Content-Type",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
}


def handler(event, context):
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}
    try:
        result = _handle(json.loads(event["body"]))
        return {"statusCode": 200, "headers": CORS, "body": json.dumps(result)}
    except Exception as e:
        return {"statusCode": 500, "headers": CORS, "body": json.dumps({"error": str(e)})}


def _handle(body):
    aws = body["credentials"]
    session = boto3.Session(
        aws_access_key_id=aws["AccessKeyId"],
        aws_secret_access_key=aws["SecretAccessKey"],
        aws_session_token=aws.get("SessionToken"),
        region_name=body.get("region", "eu-west-1"),
    )

    sm = session.client("secretsmanager")
    resp = sm.get_secret_value(SecretId=f"{body['customer']}/hubspot")
    creds = json.loads(resp["SecretString"])
    token, auth_type = _authenticate(creds)

    if body["action"] == "discover":
        return _discover(token, auth_type, body.get("filter", ""))

    return _query(
        token, auth_type,
        body["object"],
        body["queryType"],
        body.get("limit", 100),
        body.get("properties", []),
        body.get("filters", []),
    )


def _authenticate(creds):
    if creds.get("hapikey"):
        return creds["hapikey"], "hapikey"

    if all(k in creds for k in ("client_id", "client_secret", "refresh_token")):
        r = requests.post(
            "https://api.hubapi.com/oauth/v1/token",
            data={
                "grant_type": "refresh_token",
                "client_id": creds["client_id"],
                "client_secret": creds["client_secret"],
                "refresh_token": creds["refresh_token"],
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        return r.json()["access_token"], "bearer"

    for key in ("access_token", "token", "api_key"):
        if creds.get(key):
            return creds[key], "bearer"

    raise ValueError(f"No usable token found. Keys: {list(creds.keys())}")


def _headers(token, auth_type):
    if auth_type == "bearer":
        return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    return {"Content-Type": "application/json"}


def _params(token, auth_type, extra=None):
    p = {"hapikey": token} if auth_type == "hapikey" else {}
    if extra:
        p.update(extra)
    return p


def _get(token, auth_type, path, params=None):
    r = requests.get(
        f"{HUBSPOT_BASE}{path}",
        headers=_headers(token, auth_type),
        params=_params(token, auth_type, params),
        timeout=REQUEST_TIMEOUT,
    )
    r.raise_for_status()
    return r.json()


def _post(token, auth_type, path, payload):
    r = requests.post(
        f"{HUBSPOT_BASE}{path}",
        headers=_headers(token, auth_type),
        params=_params(token, auth_type),
        json=payload,
        timeout=REQUEST_TIMEOUT,
    )
    r.raise_for_status()
    return r.json()


def _count(token, auth_type, obj):
    try:
        d = _post(token, auth_type, f"/crm/v3/objects/{obj}/search",
                  {"filterGroups": [], "limit": 1, "properties": ["hs_object_id"]})
        return d.get("total", 0)
    except Exception:
        return -1


def _flatten(record):
    flat = {"id": record.get("id", "")}
    flat.update(record.get("properties", {}))
    return flat


def _discover(token, auth_type, filter_term):
    schemas = []
    try:
        schemas = _get(token, auth_type, "/crm/v3/schemas").get("results", [])
    except Exception:
        pass

    all_objects = [{"name": o, "label": o.title(), "type": "standard"} for o in HUBSPOT_STANDARD_OBJECTS]
    for s in schemas:
        all_objects.append({
            "name": s.get("fullyQualifiedName", s.get("name", "")),
            "label": s.get("labels", {}).get("singular", s.get("name", "")),
            "type": "custom",
        })

    if filter_term:
        term = filter_term.lower()
        all_objects = [o for o in all_objects if term in o["name"].lower() or term in o["label"].lower()]

    rows = []
    for obj in all_objects:
        count = _count(token, auth_type, obj["name"])
        rows.append({**obj, "record_count": count if count >= 0 else "Error"})
        time.sleep(DISCOVER_DELAY)

    return {"type": "discover", "rows": rows}


def _query(token, auth_type, obj, qtype, limit, props, filters):
    if qtype == "count":
        return {"type": "count", "total": _count(token, auth_type, obj)}

    if qtype == "list":
        params = {"limit": min(limit, 100)}
        if props:
            params["properties"] = ",".join(props)
        records = _get(token, auth_type, f"/crm/v3/objects/{obj}", params).get("results", [])
        return {"type": "list", "rows": [_flatten(r) for r in records]}

    if qtype == "shape":
        all_props = _get(token, auth_type, f"/crm/v3/properties/{obj}").get("results", [])
        rows = [
            {"name": p["name"], "label": p["label"], "type": p["type"],
             "fieldType": p["fieldType"], "group": p["groupName"]}
            for p in all_props
        ]
        return {"type": "shape", "rows": rows, "excel": _make_excel(rows, "Object Shape"),
                "filename": f"{obj}_shape.xlsx"}

    if qtype == "all":
        all_props = _get(token, auth_type, f"/crm/v3/properties/{obj}").get("results", [])
        prop_names = [p["name"] for p in all_props]
        records, after = [], None
        while len(records) < limit:
            batch_size = min(100, limit - len(records))
            payload = {"filterGroups": [], "properties": prop_names, "limit": batch_size}
            if after:
                payload["after"] = after
            data = _post(token, auth_type, f"/crm/v3/objects/{obj}/search", payload)
            batch = data.get("results", [])
            records.extend(batch)
            after = data.get("paging", {}).get("next", {}).get("after")
            if not after or not batch:
                break
        rows = [_flatten(r) for r in records]
        return {"type": "all", "rows": rows, "excel": _make_excel(rows, "Query Results"),
                "filename": f"{obj}_records.xlsx"}

    if qtype == "search":
        payload = {
            "filterGroups": [{"filters": filters}] if filters else [],
            "limit": min(limit, 100),
        }
        if props:
            payload["properties"] = props
        result = _post(token, auth_type, f"/crm/v3/objects/{obj}/search", payload)
        records = result.get("results", [])
        rows = [_flatten(r) for r in records]
        return {"type": "search", "rows": rows, "total": result.get("total", len(records))}

    raise ValueError(f"Unknown query type: {qtype}")


def _make_excel(records, sheet_name="Results"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not records:
        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()

    all_fields = list(dict.fromkeys(k for r in records for k in r.keys()))
    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")

    for ci, field in enumerate(all_fields, 1):
        cell = ws.cell(row=1, column=ci, value=field)
        cell.fill = fill
        cell.font = font

    for ri, row in enumerate(records, 2):
        for ci, field in enumerate(all_fields, 1):
            val = row.get(field, "")
            ws.cell(row=ri, column=ci, value=str(val) if isinstance(val, dict) else val)

    for col in ws.columns:
        width = min(max((len(str(c.value or "")) for c in col), default=10) + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()
