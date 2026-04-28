import base64
import io
import json

import boto3
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REQUEST_TIMEOUT = 25

CORS = {
    "Content-Type": "application/json",
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Headers": "Content-Type",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
}


def handler(event, context):
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}
    try:
        result = _handle(json.loads(event["body"]))
        return {"statusCode": 200, "headers": CORS, "body": json.dumps(result)}
    except Exception as e:
        return {"statusCode": 500, "headers": CORS, "body": json.dumps({"error": str(e)})}


def _handle(body):
    aws = body["credentials"]
    session = boto3.Session(
        aws_access_key_id=aws["AccessKeyId"],
        aws_secret_access_key=aws["SecretAccessKey"],
        aws_session_token=aws.get("SessionToken"),
        region_name=body.get("region", "eu-west-1"),
    )

    sm = session.client("secretsmanager")
    resp = sm.get_secret_value(SecretId=f"{body['customer']}/salesforce")
    creds = json.loads(resp["SecretString"])

    instance_url = creds.get("instance_url", "").rstrip("/")
    if not instance_url:
        raise ValueError("instance_url not found in credentials")

    token = _get_token(creds, instance_url)

    if body["action"] == "discover":
        return _discover(token, instance_url, body.get("filter", ""))

    return _query(
        token, instance_url,
        body.get("object", ""),
        body["queryType"],
        body.get("limit", 10),
        body.get("soql"),
    )


def _get_token(creds, instance_url):
    token_url = f"{instance_url}/services/oauth2/token"

    if "refresh_token" in creds:
        try:
            r = requests.post(token_url, data={
                "grant_type": "refresh_token",
                "client_id": creds["client_id"],
                "client_secret": creds["client_secret"],
                "refresh_token": creds["refresh_token"],
            }, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            return r.json()["access_token"]
        except Exception:
            pass

    if "username" in creds and "password" in creds:
        try:
            r = requests.post(token_url, data={
                "grant_type": "password",
                "client_id": creds["client_id"],
                "client_secret": creds["client_secret"],
                "username": creds["username"],
                "password": creds["password"] + creds.get("security_token", ""),
            }, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            return r.json()["access_token"]
        except Exception:
            pass

    r = requests.post(token_url, data={
        "grant_type": "client_credentials",
        "client_id": creds["client_id"],
        "client_secret": creds["client_secret"],
    }, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    return r.json()["access_token"]


def _headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def _sf_query(token, instance_url, soql):
    r = requests.get(
        f"{instance_url}/services/data/v59.0/query",
        headers=_headers(token),
        params={"q": soql},
        timeout=REQUEST_TIMEOUT,
    )
    r.raise_for_status()
    return r.json()


def _discover(token, instance_url, filter_term):
    r = requests.get(
        f"{instance_url}/services/data/v59.0/sobjects",
        headers=_headers(token),
        timeout=REQUEST_TIMEOUT,
    )
    r.raise_for_status()
    all_objects = r.json().get("sobjects", [])

    if filter_term:
        term = filter_term.lower()
        all_objects = [o for o in all_objects if term in o["name"].lower() or term in o.get("label", "").lower()]

    rows = []
    for obj in all_objects:
        is_queryable = obj.get("queryable", False)
        if is_queryable:
            try:
                result = _sf_query(token, instance_url, f"SELECT COUNT() FROM {obj['name']}")
                count = result.get("totalSize", 0)
            except Exception:
                count = "Error"
        else:
            count = "N/A"
        rows.append({
            "name": obj["name"],
            "label": obj.get("label", ""),
            "queryable": is_queryable,
            "record_count": count,
        })

    return {"type": "discover", "rows": rows}


def _query(token, instance_url, obj, qtype, limit, soql):
    if qtype == "shape":
        r = requests.get(
            f"{instance_url}/services/data/v59.0/sobjects/{obj}/describe",
            headers=_headers(token),
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        fields = [
            {"name": f["name"], "label": f["label"], "type": f["type"], "length": f["length"]}
            for f in r.json().get("fields", [])
        ]
        return {"type": "shape", "rows": fields, "excel": _make_excel(fields, "Object Shape"),
                "filename": f"{obj}_shape.xlsx"}

    if soql:
        query_str = soql
    elif qtype == "count":
        query_str = f"SELECT COUNT() FROM {obj}"
    elif qtype == "list":
        query_str = f"SELECT Id, Name FROM {obj} LIMIT 20"
    elif qtype == "all":
        query_str = f"SELECT FIELDS(ALL) FROM {obj} LIMIT {limit}"
    else:
        query_str = f"SELECT Id FROM {obj} LIMIT 10"

    result = _sf_query(token, instance_url, query_str)
    records = result.get("records", [])
    total = result.get("totalSize", 0)

    if qtype == "count":
        count_val = records[0].get("expr0", total) if records else total
        return {"type": "count", "total": count_val}

    clean = [{k: v for k, v in r.items() if k != "attributes"} for r in records]

    if qtype in ("all", "custom"):
        return {
            "type": qtype, "rows": clean, "total": total,
            "excel": _make_excel(clean, "Query Results"),
            "filename": f"{obj or 'query'}_results.xlsx",
        }

    return {"type": qtype, "rows": clean, "total": total}


def _make_excel(records, sheet_name="Results"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not records:
        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()

    all_fields = list(dict.fromkeys(k for r in records for k in r.keys()))
    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")

    for ci, field in enumerate(all_fields, 1):
        cell = ws.cell(row=1, column=ci, value=field)
        cell.fill = fill
        cell.font = font

    for ri, row in enumerate(records, 2):
        for ci, field in enumerate(all_fields, 1):
            val = row.get(field, "")
            ws.cell(row=ri, column=ci, value=str(val) if isinstance(val, dict) else val)

    for col in ws.columns:
        width = min(max((len(str(c.value or "")) for c in col), default=10) + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()
